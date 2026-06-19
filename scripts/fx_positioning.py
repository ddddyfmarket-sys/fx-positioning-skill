#!/usr/bin/env python3
"""
FX Positioning Monitor — self-contained script.

Fetches CFTC TFF (Traders in Financial Futures) Futures+Options Combined data
via the CFTC Socrata API, computes speculative positioning metrics, and writes
a formatted "FX Positioning" sheet into Positioning_Data.xlsx for 11 FX markets:
    EUR, JPY, GBP, CHF, CAD, AUD, NZD, MXN, BRL, ZAR, DXY

Methodology
    Net % OI  = (Leveraged Funds + Asset Managers) net / Open Interest × 100
    Percentile >= 90 in any window → green cell  (crowded LONG)
    Percentile <= 10 in any window → red cell    (crowded SHORT)
    Negative numbers shown in red font.
    (Lookback windows: 13W ~3M tactical, 52W ~1Y cyclical, 3Y ~156-report
     structural; plus full-history reads in the CSV.)

API response is cached locally for 24 hours so repeated runs don't hit the API.

Outputs (written to --outdir, default = current working directory):
    Positioning_Data.xlsx    formatted table + all three charts embedded
    ytd_positioning.png      standalone YTD distribution box plot
    history_positioning.png  full-history small-multiples, y-axis = rolling 3Y percentile
    holdings_flows_positioning.png holdings (52W Z level) vs flows (1M change in 52W Z) scatter, one dot per ccy
    positioning_table.csv    machine-readable table incl. Lev / AstMgr split

Usage:
    python3 fx_positioning.py                      # outputs to CWD
    python3 fx_positioning.py --outdir /path/out   # outputs to given dir
    python3 fx_positioning.py --refresh            # force re-fetch from API
"""

import io
import json
import sys
from collections import defaultdict
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

import openpyxl
import requests
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ── Config ───────────────────────────────────────────────────────────────────

EXCEL_NAME     = "Positioning_Data.xlsx"
SHEET_NAME     = "FX Positioning"
DATA_START_ROW = 6
N_COLS         = 17


def resolve_outdir() -> Path:
    """Output directory: --outdir <path> if given, else the current working
    directory. The skill runs this from the PM's working folder, so outputs
    land where they're invoked rather than inside the skill package."""
    outdir = Path.cwd()
    if "--outdir" in sys.argv:
        i = sys.argv.index("--outdir")
        if i + 1 < len(sys.argv):
            outdir = Path(sys.argv[i + 1]).expanduser()
    outdir.mkdir(parents=True, exist_ok=True)
    return outdir

CACHE_DIR  = Path.home() / ".fx_cache"
CACHE_DIR.mkdir(exist_ok=True)
API_CACHE  = CACHE_DIR / "tff_combined_api.json"

SOCRATA_URL = "https://publicreporting.cftc.gov/resource/yw9f-hn96.json"
FETCH_LIMIT = 50000   # max rows per request; full FX dataset is ~11k rows

CCY_ORDER = ["EUR", "JPY", "GBP", "CHF", "CAD", "AUD", "NZD", "MXN", "BRL", "ZAR", "DXY"]

# Per-currency chart display start date (rolling 3Y percentile still computed over the
# full underlying series; only the displayed date window is clipped)
CHART_STARTS = {
    "ZAR": date(2016, 1, 1),   # pre-2016 ZAR data is very sparse (<5 obs/year)
}

# Flat market-name → CCY lookup covering all known CFTC name variants.
# CFTC renamed several markets in 2022; both old and new names are included.
MARKET_NAMES = {
    # EUR
    "EURO FX - CHICAGO MERCANTILE EXCHANGE":                "EUR",
    # JPY
    "JAPANESE YEN - CHICAGO MERCANTILE EXCHANGE":           "JPY",
    # GBP  (renamed 2022)
    "BRITISH POUND - CHICAGO MERCANTILE EXCHANGE":          "GBP",
    "BRITISH POUND STERLING - CHICAGO MERCANTILE EXCHANGE": "GBP",
    # CHF
    "SWISS FRANC - CHICAGO MERCANTILE EXCHANGE":            "CHF",
    # CAD
    "CANADIAN DOLLAR - CHICAGO MERCANTILE EXCHANGE":        "CAD",
    # AUD
    "AUSTRALIAN DOLLAR - CHICAGO MERCANTILE EXCHANGE":      "AUD",
    # NZD  (renamed 2022)
    "NZ DOLLAR - CHICAGO MERCANTILE EXCHANGE":              "NZD",
    "NEW ZEALAND DOLLAR - CHICAGO MERCANTILE EXCHANGE":     "NZD",
    # MXN
    "MEXICAN PESO - CHICAGO MERCANTILE EXCHANGE":           "MXN",
    # BRL
    "BRAZILIAN REAL - CHICAGO MERCANTILE EXCHANGE":         "BRL",
    # ZAR  (renamed 2022)
    "SO AFRICAN RAND - CHICAGO MERCANTILE EXCHANGE":        "ZAR",
    "SOUTH AFRICAN RAND - CHICAGO MERCANTILE EXCHANGE":     "ZAR",
    # DXY  (renamed 2022)
    "USD INDEX - ICE FUTURES U.S.":                         "DXY",
    "U.S. DOLLAR INDEX - ICE FUTURES U.S.":                "DXY",
}

# Socrata field names for the columns we need
_API_FIELDS = ",".join([
    "market_and_exchange_names",
    "report_date_as_yyyy_mm_dd",
    "open_interest_all",
    "lev_money_positions_long",
    "lev_money_positions_short",
    "asset_mgr_positions_long",
    "asset_mgr_positions_short",
])

_MKT_FILTER = " OR ".join(
    f"market_and_exchange_names='{m}'" for m in MARKET_NAMES
)


# ── CFTC Socrata API fetch ────────────────────────────────────────────────────

def _cache_fresh() -> bool:
    """True if the local cache exists and is less than 24 hours old."""
    if not API_CACHE.exists():
        return False
    age = datetime.now() - datetime.fromtimestamp(API_CACHE.stat().st_mtime)
    return age < timedelta(hours=24)


def fetch_all(force_refresh: bool = False) -> dict:
    """
    Fetch full CFTC TFF Combined history (2006–present) via Socrata API.
    Caches the raw response for 24 hours; use force_refresh=True to bypass.
    Returns dict keyed by CCY: {'latest': {...}, 'prev': {...}, 'history': [...]}.
    """
    if not force_refresh and _cache_fresh():
        raw = json.loads(API_CACHE.read_text())
        print(f"  Using cached data ({len(raw)} rows, "
              f"updated {datetime.fromtimestamp(API_CACHE.stat().st_mtime):%Y-%m-%d %H:%M})")
    else:
        print("  Fetching from CFTC Socrata API...")
        raw, offset = [], 0
        while True:
            resp = requests.get(SOCRATA_URL, params={
                "$where":  f"({_MKT_FILTER}) AND futonly_or_combined='Combined'",
                "$select": _API_FIELDS,
                "$order":  "report_date_as_yyyy_mm_dd ASC",
                "$limit":  FETCH_LIMIT,
                "$offset": offset,
            }, timeout=60)
            resp.raise_for_status()
            batch = resp.json()
            raw.extend(batch)
            if len(batch) < FETCH_LIMIT:
                break
            offset += FETCH_LIMIT
        API_CACHE.write_text(json.dumps(raw))
        print(f"  Fetched {len(raw)} rows → cached to {API_CACHE.name}")

    # Parse raw API rows into per-CCY history.
    # Keyed by date so a duplicate report date (re-published revisions) keeps
    # only the last-seen row rather than appending a duplicate observation.
    by_ccy: dict[str, dict[date, dict]] = defaultdict(dict)
    for row in raw:
        # Map the CFTC market name (old or new variant) to our ticker; skip
        # any market we don't track.
        ccy = MARKET_NAMES.get(row.get("market_and_exchange_names", ""))
        if ccy is None:
            continue
        dt  = date.fromisoformat(row["report_date_as_yyyy_mm_dd"][:10])
        oi  = int(row.get("open_interest_all") or 0)
        if oi == 0:
            continue  # avoid divide-by-zero; a 0-OI report is unusable anyway
        # Net % OI per trader group = (long − short) / open interest × 100.
        # `or 0` guards against null/missing fields in the API response.
        lev_net = (int(row.get("lev_money_positions_long") or 0)
                 - int(row.get("lev_money_positions_short") or 0)) / oi * 100
        am_net  = (int(row.get("asset_mgr_positions_long") or 0)
                 - int(row.get("asset_mgr_positions_short") or 0)) / oi * 100
        by_ccy[ccy][dt] = {
            "date":          dt,
            "open_int":      oi,
            "lev_net_pct":   lev_net,
            "am_net_pct":    am_net,
            "total_net_pct": lev_net + am_net,
        }

    # Build the final structure: chronological history plus quick handles on
    # the two most recent reports (latest + prior week for WoW change).
    result = {}
    for ccy, date_map in by_ccy.items():
        rows = sorted(date_map.values(), key=lambda x: x["date"])
        if len(rows) < 2:
            continue  # need at least two points to compute a week-over-week change
        result[ccy] = {"latest": rows[-1], "prev": rows[-2], "history": rows}

    total_obs = sum(len(v["history"]) for v in result.values())
    print(f"  Loaded {total_obs} observations across {len(result)} currencies.")
    return result


# ── Math helpers ──────────────────────────────────────────────────────────────

def pctl_rank(value: float, history: list[float]) -> float:
    """Percentile rank (0–100) of `value` within `history`, inclusive.

    Counts how many observations are <= value. `history` normally already
    contains `value` itself (it is the latest point of the window), so a
    fresh all-time high reads ~100 and an all-time low reads near 0.
    """
    n = len(history)
    if n == 0:
        return 50.0  # neutral fallback when there is nothing to rank against
    return sum(1 for v in history if v <= value) / n * 100


def z_score(value: float, history: list[float]) -> float:
    """Standard score of `value` vs `history` using the sample stdev (n−1).

    Returns 0.0 when there are too few points (n < 2) or the window is flat
    (stdev ≈ 0), which avoids a divide-by-zero.
    """
    n = len(history)
    if n < 2:
        return 0.0
    mean  = sum(history) / n
    stdev = (sum((v - mean) ** 2 for v in history) / (n - 1)) ** 0.5
    return 0.0 if stdev < 1e-10 else (value - mean) / stdev


# ── Compute positioning row ───────────────────────────────────────────────────

def compute_row(ccy: str, ccy_data: dict) -> dict | None:
    """Build one display row of positioning metrics for a single currency.

    Returns None if there is too little history (<13 weekly reports) for the
    percentile/Z columns to be meaningful.
    """
    history = ccy_data["history"]
    if len(history) < 13:
        return None

    latest = ccy_data["latest"]
    prev   = ccy_data["prev"]

    # Lookback windows of the Total Net % OI series (most recent N reports),
    # ordered by horizon: 13W ≈ 3M (tactical), 52W ≈ 1Y (cyclical), 3Y ≈ 156
    # weekly reports (structural/multi-year). The 13-week window is the shortest
    # that keeps the percentile/Z reasonably meaningful (a 4-week window was too
    # thin). Where a currency has fewer than a window's worth of reports, the
    # slice simply uses all available — pctl_rank / z_score guard small n.
    hist_13w  = [r["total_net_pct"] for r in history[-13:]]
    hist_52w  = [r["total_net_pct"] for r in history[-52:]]
    hist_3y   = [r["total_net_pct"] for r in history[-156:]]
    hist_full = [r["total_net_pct"] for r in history]   # full history (2006→); the
    #                                                     basis for the Hist Z / Hist
    #                                                     Pctl full-sample reads in the
    #                                                     CSV — NOT a fixed window.

    cur   = latest["total_net_pct"]
    wow   = cur - prev["total_net_pct"]                 # vs last week (history[-2])
    # Per-group week-over-week change, so the Lev / AstMgr legs of the WoW move
    # can be read alongside the aggregate (they need not move the same way).
    lev_wow = latest["lev_net_pct"] - prev["lev_net_pct"]
    am_wow  = latest["am_net_pct"]  - prev["am_net_pct"]
    # Month-over-month ≈ change vs 4 reports ago (history[-5] = current minus 4 weeks).
    has_mom = len(history) >= 5
    mom     = round(cur - history[-5]["total_net_pct"], 2) if has_mom else None
    lev_mom = round(latest["lev_net_pct"] - history[-5]["lev_net_pct"], 2) if has_mom else None
    am_mom  = round(latest["am_net_pct"]  - history[-5]["am_net_pct"],  2) if has_mom else None
    p52_r = round(pctl_rank(cur, hist_52w), 1)

    return {
        "CCY":        ccy,
        "13W Pctl":   round(pctl_rank(cur, hist_13w), 1),
        "13W Z":      round(z_score(cur, hist_13w), 2),
        "52W Pctl":   p52_r,
        "52W Z":      round(z_score(cur, hist_52w), 2),
        "3Y Pctl":    round(pctl_rank(cur, hist_3y), 1),
        "3Y Z":       round(z_score(cur, hist_3y), 2),
        # Full-history reads (full sample, 2006→) — the correct basis for any
        # "historic extreme" statement. These can diverge sharply from the 52W
        # reads: a position can sit at the top of its trailing year (52W Pctl ~98)
        # yet be mid- or low-range on full history (Hist Pctl ~20) — e.g. a
        # heavily-covered but still-net-short book. (The history chart's y-axis is
        # the rolling 3Y percentile, a separate point-in-time measure.)
        "Hist Z":     round(z_score(cur, hist_full), 2),
        "Hist Pctl":  round(pctl_rank(cur, hist_full), 1),
        # WoW Δ and MoM Δ split by trader group (Total = Lev + AstMgr), so each
        # change group mirrors the NET POSITION group's Total/Lev/AstMgr layout.
        "WoW Chg":    round(wow, 2),
        "WoW Lev":    round(lev_wow, 2),
        "WoW AstMgr": round(am_wow, 2),
        "MoM Chg":    mom,
        "MoM Lev":    lev_mom,
        "MoM AstMgr": am_mom,
        "Total Net":  round(cur, 2),
        "LevFd Net":  round(latest["lev_net_pct"], 2),
        "AstMgr Net": round(latest["am_net_pct"], 2),
        "Open Int":   latest["open_int"],
        "Date":       latest["date"],
    }


# ── Cell styles ───────────────────────────────────────────────────────────────

_PCTL_HIGH_FILL = PatternFill("solid", fgColor="C6EFCE")
_PCTL_HIGH_FONT = Font(color="276221", bold=True)
_PCTL_LOW_FILL  = PatternFill("solid", fgColor="FFC7CE")
_PCTL_LOW_FONT  = Font(color="9C0006", bold=True)
_NEG_FONT       = Font(color="9C0006")

_HDR_GRP_FILL = PatternFill("solid", fgColor="1F3864")
_HDR_GRP_FONT = Font(bold=True, color="FFFFFF", size=8)
_HDR_GRP_ALGN = Alignment(horizontal="center", vertical="center")
_HDR_SUB_FILL = PatternFill("solid", fgColor="2F5496")
_HDR_SUB_FONT = Font(bold=True, color="FFFFFF", size=8)
_HDR_SUB_ALGN = Alignment(horizontal="center", vertical="center")

_THIN     = Side(style="thin", color="D9D9D9")
_BORDER   = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_ALT_FILL = PatternFill("solid", fgColor="F5F8FD")

# Col:  1=CCY  2=13W Pctl  3=13W Z  4=52W Pctl  5=52W Z  6=3Y Pctl  7=3Y Z
#       8=WoW Total   9=WoW Lev   10=WoW AstMgr
#       11=MoM Total  12=MoM Lev  13=MoM AstMgr
#       14=Total Net  15=Lev  16=AstMgr  17=OI
_NUMFMT = {
    2: "0.0", 3: "0.00", 4: "0.0", 5: "0.00", 6: "0.0", 7: "0.00",
    8: "+0.00;-0.00",  9: "+0.00;-0.00", 10: "+0.00;-0.00",
    11: "+0.00;-0.00", 12: "+0.00;-0.00", 13: "+0.00;-0.00",
    14: "0.00", 15: "0.00", 16: "0.00", 17: "#,##0",
}
_COL_WIDTHS = [10, 7, 7.5, 7, 7.5, 7, 7.5, 8, 8, 8, 8, 8, 8, 8, 8, 8.5, 11]
_PCTL_COLS  = {2, 4, 6}                              # 13W / 52W / 3Y percentile columns
_NEG_COLS   = {3, 5, 7, 8, 9, 10, 11, 12, 13, 14, 15, 16}  # Z, change & net cols (red if < 0)


def _safe_merge(ws, r1, c1, r2, c2):
    ref = f"{get_column_letter(c1)}{r1}:{get_column_letter(c2)}{r2}"
    try:
        ws.merge_cells(ref)
    except Exception:
        ws.unmerge_cells(ref)
        ws.merge_cells(ref)


def _write_headers(ws) -> None:
    for c in range(1, N_COLS + 1):
        ws.cell(4, c).fill = _HDR_GRP_FILL

    for c1, c2, label in [
        (1,  1,  None),
        (2,  3,  "13W (% OI)"),
        (4,  5,  "52W (% OI)"),
        (6,  7,  "3Y (% OI)"),
        (8,  10, "WoW Δ (% OI)"),
        (11, 13, "MoM Δ (% OI)"),
        (14, 16, "NET POSITION (% OI)"),
        (17, 17, None),
    ]:
        if c1 < c2:
            _safe_merge(ws, 4, c1, 4, c2)
        cell = ws.cell(4, c1)
        cell.value     = label
        cell.alignment = _HDR_GRP_ALGN
        cell.fill      = _HDR_GRP_FILL
        if label:
            cell.font = _HDR_GRP_FONT

    for col, label in [
        (1, "CCY"),
        (2, "Pctl"), (3, "Z-Score"),
        (4, "Pctl"), (5, "Z-Score"),
        (6, "Pctl"), (7, "Z-Score"),
        (8, "Total"),  (9, "Lev"),  (10, "AstMgr"),
        (11, "Total"), (12, "Lev"), (13, "AstMgr"),
        (14, "Total"), (15, "Lev"), (16, "AstMgr"),
        (17, "OI"),
    ]:
        c = ws.cell(5, col)
        c.value, c.fill, c.font, c.alignment = label, _HDR_SUB_FILL, _HDR_SUB_FONT, _HDR_SUB_ALGN

    for row in (4, 5):
        for col in range(1, N_COLS + 1):
            ws.cell(row, col).border = _BORDER


def _write_data_row(ws, row_num: int, r: dict, zebra: bool) -> None:
    vals = [
        r["CCY"], r["13W Pctl"], r["13W Z"], r["52W Pctl"], r["52W Z"],
        r["3Y Pctl"], r["3Y Z"],
        r["WoW Chg"], r["WoW Lev"], r["WoW AstMgr"],
        r["MoM Chg"], r["MoM Lev"], r["MoM AstMgr"],
        r["Total Net"], r["LevFd Net"], r["AstMgr Net"],
        r["Open Int"],
    ]
    for col, val in enumerate(vals, start=1):
        c = ws.cell(row_num, col)
        c.value     = val
        c.border    = _BORDER
        c.alignment = Alignment(horizontal="left" if col == 1 else "center")
        if col in _NUMFMT and val is not None:
            c.number_format = _NUMFMT[col]
        if zebra:
            c.fill = _ALT_FILL

        # Percentile columns get a crowding highlight; everything else just
        # gets red font when negative. These two branches are mutually
        # exclusive (percentile cols are never recoloured for sign).
        if col in _PCTL_COLS and val is not None:
            if val >= 90:
                c.fill, c.font = _PCTL_HIGH_FILL, _PCTL_HIGH_FONT   # crowded long
            elif val <= 10:
                c.fill, c.font = _PCTL_LOW_FILL, _PCTL_LOW_FONT     # crowded short
        elif col in _NEG_COLS and val is not None and isinstance(val, (int, float)) and val < 0:
            c.font = _NEG_FONT


# ── YTD box plot ──────────────────────────────────────────────────────────────

def _ytd_z_scores(ccy_data: dict) -> list[dict]:
    """52-week trailing Z-score for every weekly report in the current year.

    For each YTD week we recompute the Z-score against its own trailing
    52-report window (i-51 .. i inclusive), so the last entry here matches the
    "52W Z" shown in the table. Used to draw the YTD distribution box plot.
    """
    history = ccy_data["history"]
    yr, out = datetime.now().year, []
    for i, wk in enumerate(history):
        if wk["date"].year != yr:
            continue
        window = [h["total_net_pct"] for h in history[max(0, i - 51): i + 1]]
        out.append({"date": wk["date"],
                    "score": z_score(wk["total_net_pct"], window) if len(window) >= 2 else 0.0})
    return out


def _make_ytd_chart(fx_rows: list[dict], fx_data: dict):
    try:
        import numpy as np
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        from matplotlib.lines import Line2D
    except Exception as e:
        print(f"  YTD chart skipped: {e}")
        return None

    # Order currencies most-bullish → most-bearish by current 52W Z-score.
    rows = sorted(fx_rows, key=lambda r: r["52W Z"], reverse=True)
    labels, box, cur, prev = [], [], [], []
    for r in rows:
        ytd    = _ytd_z_scores(fx_data[r["CCY"]])
        scores = [s["score"] for s in ytd] or [r["52W Z"]]  # fallback: 1 point if no YTD data
        labels.append(r["CCY"])
        box.append(scores)                                   # YTD distribution → box
        cur.append(r["52W Z"])                               # current → "x" marker
        prev.append(ytd[-2]["score"] if len(ytd) >= 2 else None)  # 1 week ago → "o" marker

    n   = len(labels)
    fig, ax = plt.subplots(figsize=(max(11, n * 0.9 + 2), 4.3))
    bp = ax.boxplot(
        box, positions=np.arange(n), widths=0.45, patch_artist=True, showfliers=False,
        medianprops=dict(color="#C55A11", linewidth=1.8),
        boxprops=dict(edgecolor="#2F5496", linewidth=1.2),
        whiskerprops=dict(color="#2F5496", linewidth=1.0, linestyle="--"),
        capprops=dict(color="#2F5496", linewidth=1.2),
    )
    for patch, score in zip(bp["boxes"], cur):
        patch.set_facecolor("#D9E5F5" if score > 0.5 else "#FFE0E0" if score < -0.5 else "#F5F5F5")

    for i, score in enumerate(cur):
        ax.plot(i, score, marker="x", markersize=12, markeredgewidth=2.6,
                color="#2F5496" if score >= 0 else "#C00000", zorder=6)
    for i, score in enumerate(prev):
        if score is not None:
            ax.plot(i, score, marker="o", markersize=5, color="#888888", zorder=5, alpha=0.8)

    ax.axhline(0,    color="#444444", linewidth=0.9)
    ax.axhline(1.5,  color="#2F5496", linewidth=0.5, linestyle=":", alpha=0.5)
    ax.axhline(-1.5, color="#C00000", linewidth=0.5, linestyle=":", alpha=0.5)
    ax.set_xticks(np.arange(n))
    ax.set_xticklabels(labels, fontsize=9, fontweight="bold", color="#1F3864")
    ax.set_ylabel("Positioning Score (52W Z-score)", fontsize=8, color="#444444")

    cot_date = max(r["Date"] for r in fx_rows)
    ax.set_title(
        f"FX Positioning Score — YTD {datetime.now().year} Distribution  "
        f"|  COT as of {cot_date}  |  {len(box[0])} weekly obs YTD",
        fontsize=10, fontweight="bold", color="#1F3864", pad=8,
    )
    ax.grid(axis="y", linestyle="--", color="#CCCCCC", alpha=0.5)
    for sp in ("top", "right"):
        ax.spines[sp].set_visible(False)
    ax.tick_params(axis="both", labelsize=8, colors="#444444")
    ax.legend(handles=[
        Line2D([0], [0], marker="x", color="#2F5496", markersize=9, markeredgewidth=2.5,
               linewidth=0, label="Current (net long)"),
        Line2D([0], [0], marker="x", color="#C00000", markersize=9, markeredgewidth=2.5,
               linewidth=0, label="Current (net short)"),
        Line2D([0], [0], marker="o", color="#888888", markersize=5,
               linewidth=0, label="1 week ago"),
    ], loc="upper right", fontsize=7, framealpha=0.85, ncol=3)
    ax.text(0.99, -0.12,
            "CFTC TFF Combined — (Leveraged Funds + Asset Managers) Net % of Open Interest.",
            transform=ax.transAxes, fontsize=6, color="#999999", ha="right")

    fig.tight_layout(rect=[0, 0.04, 1, 1])
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=130, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    buf.seek(0)
    return buf


# ── Full-history time-series chart on a rolling-3Y percentile y-axis ──────────

def _make_history_chart(fx_data: dict):
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import matplotlib.dates as mdates
    except Exception as e:
        print(f"  History chart skipped: {e}")
        return None

    n, ncols = len(CCY_ORDER), 3
    nrows    = (n + ncols - 1) // ncols

    fig, axes = plt.subplots(nrows, ncols,
                             figsize=(ncols * 5.5, nrows * 3.0 + 0.8),
                             squeeze=False)
    fig.suptitle(
        "FX Positioning — Full History  (rolling 3Y percentile of Total Net % OI: "
        "Leveraged Funds + Asset Managers)",
        fontsize=11, fontweight="bold", color="#1F3864", y=0.99,
    )

    for idx, ccy in enumerate(CCY_ORDER):
        ax      = axes[idx // ncols][idx % ncols]
        history = fx_data.get(ccy, {}).get("history", [])

        if len(history) < 2:
            ax.set_title(ccy, fontsize=9, fontweight="bold", color="#1F3864")
            ax.text(0.5, 0.5, "Insufficient data", transform=ax.transAxes,
                    ha="center", va="center", fontsize=8, color="#999999")
            ax.axis("off")
            continue

        dates  = [h["date"] for h in history]
        values = [h["total_net_pct"] for h in history]

        # Y-axis = ROLLING 3-year percentile (0–100): each observation is ranked
        # only against the trailing 156 reports (≈3Y) up to and including itself,
        # so the line reads as how stretched positioning was AT THAT TIME vs its
        # own recent-3Y range — not vs the full sample (which would use future
        # data and flatten point-in-time extremes). 50 = the 3Y median; 90/10
        # mark the long/short ends of the trailing-3Y range. Early points (before
        # 156 reports exist) rank against all history available so far.
        ROLL_3Y = 156
        pcts = [pctl_rank(values[i], values[max(0, i - ROLL_3Y + 1): i + 1])
                for i in range(len(values))]

        # Clip display window for sparse-early-data currencies
        chart_start  = CHART_STARTS.get(ccy, dates[0])
        disp_dates   = [d for d, p in zip(dates, pcts) if d >= chart_start]
        disp_pcts    = [p for d, p in zip(dates, pcts) if d >= chart_start]
        if len(disp_dates) < 2:
            disp_dates, disp_pcts = dates, pcts

        ax.plot(disp_dates, disp_pcts, color="#2F5496", linewidth=1.0, zorder=3)
        # Shade vs the median (50th pctl): above = longer than its own history,
        # below = shorter.
        ax.fill_between(disp_dates, disp_pcts, 50,
                        where=[p >= 50 for p in disp_pcts],
                        alpha=0.18, color="#276221", interpolate=True, zorder=2)
        ax.fill_between(disp_dates, disp_pcts, 50,
                        where=[p < 50 for p in disp_pcts],
                        alpha=0.18, color="#9C0006", interpolate=True, zorder=2)

        # Range-extreme bands at the 90th / 10th percentile (long / short ends).
        ax.axhline(90, color="#276221", linewidth=0.9, linestyle="--", alpha=0.75)
        ax.axhline(10, color="#9C0006",  linewidth=0.9, linestyle="--", alpha=0.75)
        ax.axhline(50, color="#888888",  linewidth=0.7, linestyle=":",  alpha=0.6)

        # Annotate band lines on the right edge
        xmax = disp_dates[-1]
        ax.annotate("90th", xy=(xmax, 90), fontsize=5.5, color="#276221",
                    va="center", ha="left", xytext=(3, 0), textcoords="offset points")
        ax.annotate("10th", xy=(xmax, 10), fontsize=5.5, color="#9C0006",
                    va="center", ha="left", xytext=(3, 0), textcoords="offset points")

        ax.set_xlim(disp_dates[0], date.today())
        ax.set_ylim(0, 100)
        ax.set_title(ccy, fontsize=9, fontweight="bold", color="#1F3864")
        ax.tick_params(axis="both", labelsize=6.5, colors="#444444")
        ax.xaxis.set_major_formatter(mdates.DateFormatter("%Y"))
        ax.xaxis.set_major_locator(mdates.YearLocator(2))
        ax.tick_params(axis="x", rotation=45)
        ax.set_yticks([0, 25, 50, 75, 100])
        ax.set_ylabel("3Y Pctl", fontsize=6, color="#444444")
        ax.grid(axis="y", linestyle="--", color="#CCCCCC", alpha=0.45)
        for sp in ("top", "right"):
            ax.spines[sp].set_visible(False)

    for idx in range(n, nrows * ncols):
        axes[idx // ncols][idx % ncols].axis("off")

    fig.text(0.99, 0.005,
             "Source: CFTC TFF Combined — rolling 3Y (156-report) percentile of (Leveraged Funds + "
             "Asset Managers) Net % of Open Interest. 50th = 3Y median; 90th/10th = range ends.",
             fontsize=6, color="#999999", ha="right")
    fig.tight_layout(rect=[0, 0.015, 1, 0.97])

    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=130, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    buf.seek(0)
    return buf


# ── Holdings-vs-flows scatter: 52W Z level vs 1-month change in 52W Z-score ────

def _z52_at(values: list[float], end_idx: int) -> float:
    """52W Z-score of the observation at position `end_idx`, computed on its own
    trailing 52-report window (end_idx-51 .. end_idx inclusive). Lets us read the
    Z-score as it stood N weeks ago, so a change in Z is measured on the matching
    shifted window rather than against today's window."""
    window = values[max(0, end_idx - 51): end_idx + 1]
    return z_score(values[end_idx], window) if len(window) >= 2 else 0.0


def _make_momentum_scatter(fx_data: dict, lag: int = 4):
    """Scatter of *holdings* vs *flows*, one dot per currency:

        y = holdings: current 52W Z-score (how stretched the position is on the year)
        x = flows: change in that 52W Z-score over the last `lag` reports (~1 month)

    Recomputes each point's prior 52W Z on its own shifted window, so the x-axis
    is a true change in the standardized score, not a raw %OI delta. The four
    quadrants read as long/short (above/below 0) × adding/paring (right/left):
    top-right = long & extending, bottom-right = short & covering, etc. This is a
    supplement to the YTD/history charts — the cross-sectional 'where is each
    book, and which way is it moving' snapshot."""
    try:
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
    except Exception as e:
        print(f"  Holdings-vs-flows scatter skipped: {e}")
        return None

    pts = []  # (ccy, x=ΔZ over lag, y=level Z)
    for ccy in CCY_ORDER:
        history = fx_data.get(ccy, {}).get("history", [])
        vals = [r["total_net_pct"] for r in history]
        if len(vals) < 52 + lag + 1:
            continue  # need a full 52W window both now and `lag` reports ago
        z_now = _z52_at(vals, len(vals) - 1)
        z_lag = _z52_at(vals, len(vals) - 1 - lag)
        pts.append((ccy, z_now - z_lag, z_now))

    if not pts:
        print("  Holdings-vs-flows scatter skipped: no currency has enough history")
        return None

    fig, ax = plt.subplots(figsize=(9, 7))
    xs = [p[1] for p in pts]
    ys = [p[2] for p in pts]
    ax.scatter(xs, ys, s=70, color="#2F5496", zorder=5,
               edgecolor="white", linewidth=0.7)
    for ccy, x, y in pts:                       # highlight the dollar index
        if ccy == "DXY":
            ax.scatter([x], [y], s=90, color="#27AAE1", zorder=6,
                       edgecolor="white", linewidth=0.8)

    # Per-currency label nudges to limit overlap in the central cluster.
    off = {"EUR": (8, 4), "JPY": (8, -12), "GBP": (8, 4), "CHF": (8, 4),
           "CAD": (-12, 8), "AUD": (8, 4), "NZD": (-14, -12), "MXN": (8, -4),
           "BRL": (8, 6), "ZAR": (8, -12), "DXY": (10, 6)}
    for ccy, x, y in pts:
        dx, dy = off.get(ccy, (8, 4))
        ax.annotate(ccy, (x, y), fontsize=9, fontweight="bold", color="#1F3864",
                    xytext=(dx, dy), textcoords="offset points")

    ax.axhline(0, color="#888", linewidth=1.0)
    ax.axvline(0, color="#888", linewidth=1.0)
    xpad = max(0.6, max(abs(v) for v in xs) * 1.25)
    ypad = max(0.6, max(abs(v) for v in ys) * 1.25)
    ax.set_xlim(-xpad, xpad)
    ax.set_ylim(-ypad, ypad)

    q = dict(fontsize=7.5, color="#AAAAAA", ha="center", va="center", style="italic")
    ax.text( xpad * 0.6,  ypad * 0.92, "LONG & extending", **q)
    ax.text(-xpad * 0.6,  ypad * 0.92, "LONG, paring",     **q)
    ax.text( xpad * 0.6, -ypad * 0.92, "SHORT, covering",  **q)
    ax.text(-xpad * 0.6, -ypad * 0.92, "SHORT & extending", **q)

    ax.set_xlabel(f"Flows — 1M Change in 52W Z-Score (vs {lag} reports ago)", fontsize=10)
    ax.set_ylabel("Holdings — 52W Z-Score (level)", fontsize=10)
    cot_date = max(h["date"] for v in fx_data.values() for h in [v["history"][-1]])
    ax.set_title(
        "FX Speculative Positioning: Holdings vs Flows (52W Z-Score)\n"
        f"CFTC TFF Combined — (Lev Funds + Asset Mgrs) Net %OI · COT {cot_date}",
        fontsize=11, fontweight="bold", color="#1F3864",
    )
    ax.grid(True, linestyle="--", color="#DDD", alpha=0.5, zorder=0)
    for sp in ("top", "right"):
        ax.spines[sp].set_visible(False)

    fig.tight_layout()
    buf = io.BytesIO()
    fig.savefig(buf, format="png", dpi=140, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    buf.seek(0)
    return buf


# ── Write Excel ───────────────────────────────────────────────────────────────

def write_excel(fx_rows: list[dict], fx_data: dict, cot_date, outdir: Path) -> None:
    excel_path = outdir / EXCEL_NAME
    wb = openpyxl.load_workbook(excel_path) if excel_path.exists() else openpyxl.Workbook()

    if SHEET_NAME in wb.sheetnames:
        del wb[SHEET_NAME]
    ws = wb.create_sheet(SHEET_NAME, 0)

    for sn in list(wb.sheetnames):
        if sn != SHEET_NAME:
            del wb[sn]

    # Row 1: title
    _safe_merge(ws, 1, 1, 1, N_COLS)
    t = ws.cell(1, 1)
    t.value     = "FX POSITIONING MONITOR  (CFTC Speculative Net, % of Open Interest)"
    t.font      = Font(bold=True, color="FFFFFF", size=16)
    t.fill      = PatternFill("solid", fgColor="1F4E78")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    # Row 2: date stamp
    hkt_now = datetime.now(timezone(timedelta(hours=8)))
    _safe_merge(ws, 2, 1, 2, N_COLS)
    s = ws.cell(2, 1)
    s.value     = (f"Report Date: {hkt_now:%Y-%m-%d %H:%M} HKT   |   COT Data: {cot_date}   |   "
                   f"Net = Leveraged Funds + Asset Managers (% OI)")
    s.font      = Font(italic=True, size=9, color="444444")
    s.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 18

    # Row 3: spacer | Rows 4–5: headers | Rows 6–16: data
    _write_headers(ws)
    by_ccy = {r["CCY"]: r for r in fx_rows}
    for i, ccy in enumerate(CCY_ORDER):
        row_num = DATA_START_ROW + i
        r = by_ccy.get(ccy)
        if r is None:
            ws.cell(row_num, 1).value = ccy
            continue
        _write_data_row(ws, row_num, r, zebra=(i % 2 == 1))

    for j, w in enumerate(_COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(j)].width = w
    ws.freeze_panes = "A6"

    # Legend
    foot_row = DATA_START_ROW + len(CCY_ORDER) + 1
    for i, (text, bold) in enumerate([
        ("POSITIONING HIGHLIGHTS  (percentile of Net % OI within each window: 13W / 52W / 3Y)", True),
        ("  Green cell: percentile ≥ 90th — at the long end of its range for that window", False),
        ("  Red cell:   percentile ≤ 10th — at the short end of its range for that window", False),
        ("NET = Leveraged Funds + Asset Managers as % of Open Interest  "
         "(Source: CFTC TFF, Futures+Options Combined)", False),
    ]):
        cell = ws.cell(foot_row + i, 1)
        cell.value = text
        cell.font  = Font(bold=bold, size=8, color="1F3864" if bold else "666666", italic=not bold)

    # Charts — embed in Excel AND save standalone PNGs so the analysis step
    # (Claude reading the skill output) can view them directly. openpyxl
    # consumes the buffer on add_image, so write the PNG bytes out first.
    from openpyxl.drawing.image import Image as XLImage
    chart_row = foot_row + 5

    buf = _make_ytd_chart(fx_rows, fx_data)
    if buf:
        (outdir / "ytd_positioning.png").write_bytes(buf.getvalue())
        buf.seek(0)
        img = XLImage(buf)
        img.anchor = f"A{chart_row}"
        ws.add_image(img)
        chart_row += 27
        print("  Saved → ytd_positioning.png")

    buf = _make_history_chart(fx_data)
    if buf:
        (outdir / "history_positioning.png").write_bytes(buf.getvalue())
        buf.seek(0)
        img = XLImage(buf)
        img.anchor = f"A{chart_row}"
        ws.add_image(img)
        chart_row += 27
        print("  Saved → history_positioning.png")

    buf = _make_momentum_scatter(fx_data)
    if buf:
        (outdir / "holdings_flows_positioning.png").write_bytes(buf.getvalue())
        buf.seek(0)
        img = XLImage(buf)
        img.anchor = "A113"   # fixed anchor — sits clear of the history chart above
        ws.add_image(img)
        print("  Saved → holdings_flows_positioning.png")

    wb.save(excel_path)
    print(f"  Saved → {excel_path.name}")


def write_table_csv(fx_rows: list[dict], cot_date, outdir: Path) -> None:
    """Machine-readable table for the analysis step. Includes the Leveraged
    Funds vs Asset Managers split (not shown in the console summary), since the
    AM-vs-lev divergence is a core part of the positioning read."""
    import csv

    cols = ["CCY", "13W Pctl", "13W Z", "52W Pctl", "52W Z", "3Y Pctl", "3Y Z",
            "Hist Z", "Hist Pctl",
            "WoW Chg", "WoW Lev", "WoW AstMgr", "MoM Chg", "MoM Lev", "MoM AstMgr",
            "Total Net", "LevFd Net", "AstMgr Net",
            "Open Int", "Date"]
    path = outdir / "positioning_table.csv"
    with path.open("w", newline="") as f:
        w = csv.writer(f)
        w.writerow([f"# FX Positioning Monitor — COT as of {cot_date}. "
                    f"Net = (Leveraged Funds + Asset Managers) long − short as % of "
                    f"Open Interest. 13W/52W/3Y Pctl & Z are vs trailing 13-week, 52-week "
                    f"and 3-year (156-report) windows — horizons from tactical to "
                    f"structural. Hist Z / Hist Pctl are vs FULL history (2006→). The "
                    f"history chart y-axis is the rolling 3Y percentile (90th/10th = range "
                    f"ends), showing how stretched positioning was at each point vs its "
                    f"trailing 3 years. Hist Z / Hist Pctl are the basis for any 'historic "
                    f"extreme' read. Windows can diverge sharply; cite the one you mean."])
        w.writerow(cols)
        for r in fx_rows:
            w.writerow([r.get(c, "") for c in cols])
    print(f"  Saved → {path.name}")


# ── Console summary ───────────────────────────────────────────────────────────

def print_summary(fx_rows: list[dict], cot_date) -> None:
    hdr = (f"{'CCY':8s} {'13W%':>5s} {'13WZ':>5s} {'52W%':>6s} {'52WZ':>6s} "
           f"{'3Y%':>5s} {'3YZ':>5s} {'WoW':>7s} {'MoM':>7s} {'Net%':>7s}")
    print(f"\n{'='*84}")
    print(f"  FX Positioning Monitor  |  COT as of {cot_date}")
    print(f"  Net = (Lev Funds + Asset Mgr) long − short, as % of Open Interest")
    print(f"{'='*84}")
    print(f"  {hdr}")
    print(f"  {'-'*len(hdr)}")
    for r in fx_rows:
        mom  = f"{r['MoM Chg']:+7.2f}" if r["MoM Chg"] is not None else "    n/a"
        flag = (" [CROWD LONG]"  if r["52W Pctl"] >= 90
                else " [CROWD SHORT]" if r["52W Pctl"] <= 10 else "")
        print(f"  {r['CCY']:8s} "
              f"{r['13W Pctl']:5.1f} {r['13W Z']:+5.2f} "
              f"{r['52W Pctl']:6.1f} {r['52W Z']:+6.2f} "
              f"{r['3Y Pctl']:5.1f} {r['3Y Z']:+5.2f} "
              f"{r['WoW Chg']:+7.2f} {mom} "
              f"{r['Total Net']:+7.2f}{flag}")
    print(f"{'='*84}\n")


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    print(f"\n{'='*60}")
    print(f"FX Positioning Monitor  |  {datetime.now():%Y-%m-%d %H:%M}")
    print(f"{'='*60}\n")

    force = "--refresh" in sys.argv
    print(f"Fetching CFTC TFF data (2006–{date.today().year})...")
    fx_data = fetch_all(force_refresh=force)
    print(f"Loaded {len(fx_data)} currencies.\n")

    fx_rows = []
    for ccy in CCY_ORDER:
        if ccy not in fx_data:
            print(f"  {ccy}: no data — skipping")
            continue
        row = compute_row(ccy, fx_data[ccy])
        if row is None:
            print(f"  {ccy}: insufficient history — skipping")
            continue
        fx_rows.append(row)

    if not fx_rows:
        print("No data to write.")
        return

    cot_date = max(r["Date"] for r in fx_rows)
    print_summary(fx_rows, cot_date)

    outdir = resolve_outdir()
    print(f"Writing outputs → {outdir}")
    write_excel(fx_rows, fx_data, cot_date, outdir)
    write_table_csv(fx_rows, cot_date, outdir)
    print(f"\nDone — {len(fx_rows)} FX markets updated.\n")


if __name__ == "__main__":
    main()
